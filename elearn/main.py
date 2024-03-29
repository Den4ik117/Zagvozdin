import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side


class Vacancy:
    """Класс для представления вакансии

    Attributes:
        name (str): Название вакансии
        salary_from (int): Нижняя граница вилки зарплаты
        salary_to (int): Верхняя граница вилки зарплаты
        salary_currency (str): Валюта оклада
        salary_average (int): Среднее значение зарплаты в рублях
        area_name (str): Территория, на которой числится вакансия
        year (int): Год публикации вакансии
    """
    currency_to_rub = {
        "AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
        "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055,
    }

    def __init__(self, vacancy):
        """Инициализирует объект Vacancy, высчитывает среднюю зарплату и производит конвертацию для целочисленных полей

        >>> vacancy_for_tests = Vacancy({'name': 'IT', 'salary_from': '200.00', 'salary_to': '240.00', 'salary_currency': 'GEL', 'area_name': 'Romania', 'published_at': '2020-09-20'})
        >>> type(vacancy_for_tests).__name__
        'Vacancy'
        >>> vacancy_for_tests.name
        'IT'
        >>> vacancy_for_tests.salary_from
        200
        >>> vacancy_for_tests.salary_to
        240
        >>> vacancy_for_tests.salary_currency
        'GEL'
        >>> vacancy_for_tests.area_name
        'Romania'
        >>> vacancy_for_tests.year
        2020

        Args:
            vacancy (dict[str, str]): Словарь вакансии, из которого инициализируются переменные объекта
        """
        self.name = vacancy['name']
        self.salary_from = int(float(vacancy['salary_from']))
        self.salary_to = int(float(vacancy['salary_to']))
        self.salary_currency = vacancy['salary_currency']
        self.salary_average = self.get_average_salary()
        self.area_name = vacancy['area_name']
        self.year = int(vacancy['published_at'][:4])

    def get_average_salary(self):
        """Метод высчитывает среднюю зарплату в рублях

        >>> vacancy_for_tests = Vacancy({'name': 'IT', 'salary_from': '200.00', 'salary_to': '240.00', 'salary_currency': 'GEL', 'area_name': 'Romania', 'published_at': '2020-09-20'})
        >>> vacancy_for_tests.get_average_salary()
        4782.8
        >>> vacancy_for_tests.salary_currency = 'RUR'
        >>> vacancy_for_tests.get_average_salary()
        220.0
        >>> vacancy_for_tests.salary_currency = 'USD'
        >>> vacancy_for_tests.get_average_salary()
        13345.2
        >>> vacancy_for_tests.salary_currency = 'KZT'
        >>> vacancy_for_tests.get_average_salary()
        28.6
        >>> vacancy_for_tests.salary_currency = 'KGS'
        >>> vacancy_for_tests.salary_from = 180
        >>> vacancy_for_tests.get_average_salary()
        159.6

        Returns:
            int: Средняя зарплата в рублях
        """
        return round(self.currency_to_rub[self.salary_currency] * (self.salary_from + self.salary_to) / 2, 1)


class DataSet:
    """Класс DataSet содержит методы по обработке данных, собирает статистику по вакансии и выводит результат в консоль

    Attributes:
        file_name (str): Название файла, может быть, в том числе, и путём, по которому расположен файл
        vacancy_name (str): Название вакансии, по которой будет собираться отдельная статистика
    """
    def __init__(self, file_name, vacancy_name):
        """Инициализирует объект DataSet

        Args:
            file_name (str): Название файла или путь до файл + название файла
            vacancy_name (str): Название вакансии для сбора статистики
        """
        self.file_name = file_name
        self.vacancy_name = vacancy_name

    @staticmethod
    def increment(dictionary, key, amount):
        """Статический метод, либо добавляет, либо создаёт новый ключ в словаре

        >>> dict_for_tests = {}
        >>> DataSet.increment(dict_for_tests, 2020, 10)
        >>> dict_for_tests
        {2020: 10}
        >>> DataSet.increment(dict_for_tests, 2020, 15)
        >>> dict_for_tests
        {2020: 25}
        >>> dict_for_tests = {}
        >>> DataSet.increment(dict_for_tests, 'key', [1])
        >>> dict_for_tests
        {'key': [1]}
        >>> DataSet.increment(dict_for_tests, 'key', [100])
        >>> dict_for_tests
        {'key': [1, 100]}
        >>> DataSet.increment(dict_for_tests, 'another', 17)
        >>> dict_for_tests
        {'key': [1, 100], 'another': 17}
        >>> DataSet.increment(dict_for_tests, 'another', 1)
        >>> dict_for_tests
        {'key': [1, 100], 'another': 18}

        Args:
            dictionary (dict): Словарь
            key (string or int): Ключ для словаря
            amount (int or list): Значение, которое нужно инициализировать, либо прибавить
        """
        if key in dictionary:
            dictionary[key] += amount
        else:
            dictionary[key] = amount

    @staticmethod
    def average(dictionary):
        """Метод считает среднее целочиленное значение в переданном словаре

        Args:
            dictionary (dict): словарь с листом целочисленных значения, среднее которых нужно посчитать

        Returns:
            dict: словарь, с посчитанными средними значениями
        """
        new_dictionary = {}
        for key, values in dictionary.items():
            new_dictionary[key] = int(sum(values) / len(values))
        return new_dictionary

    def csv_reader(self):
        """Метод читает CSV файл и возвращает итератор, по которому можно пройтись циклом

        Returns:
            Iterator[dict]: динамически сгенерированный словарь вакансии
        """
        with open(self.file_name, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            header = next(reader)
            header_length = len(header)
            for row in reader:
                if '' not in row and len(row) == header_length:
                    yield dict(zip(header, row))

    @staticmethod
    def sort_list_of_tuples_by_value(list_of_tuples, reverse=True):
        """Сортирует список таплов по последнему значению

        Args:
            list_of_tuples ((int, int)[]): Лист таплов для сортировке, где предполагается, что 1 значение - ключ, 2 значение - значение
            reverse (bool): порядок сортировка, по умолчанию True
        """
        list_of_tuples.sort(key=lambda a: a[-1], reverse=reverse)

    def get_statistic(self):
        """Собирает, генерирует и возвращает статистику по вакансиям

        Returns:
            tuple[dict[int, int], dict[int, int], dict[int, int], dict[int, int], dict[string, int], dict[string, int]]: 6 переменных, содержащих информацию о собранной статистике
        """
        salary = {}
        salary_of_vacancy_name = {}
        salary_city = {}
        count_of_vacancies = 0

        for vacancy_dictionary in self.csv_reader():
            vacancy = Vacancy(vacancy_dictionary)
            self.increment(salary, vacancy.year, [vacancy.salary_average])
            if vacancy.name.find(self.vacancy_name) != -1:
                self.increment(salary_of_vacancy_name, vacancy.year, [vacancy.salary_average])
            self.increment(salary_city, vacancy.area_name, [vacancy.salary_average])
            count_of_vacancies += 1

        vacancies_number = dict([(key, len(value)) for key, value in salary.items()])
        vacancies_number_by_name = dict([(key, len(value)) for key, value in salary_of_vacancy_name.items()])

        if not salary_of_vacancy_name:
            salary_of_vacancy_name = dict([(key, [0]) for key, value in salary.items()])
            vacancies_number_by_name = dict([(key, 0) for key, value in vacancies_number.items()])

        stats = self.average(salary)
        stats2 = self.average(salary_of_vacancy_name)
        stats3 = self.average(salary_city)

        stats4 = {}
        for year, salaries in salary_city.items():
            stats4[year] = round(len(salaries) / count_of_vacancies, 4)
        stats4 = list(filter(lambda a: a[-1] >= 0.01, [(key, value) for key, value in stats4.items()]))
        self.sort_list_of_tuples_by_value(stats4)
        stats5 = stats4.copy()
        stats4 = dict(stats4)
        stats3 = list(filter(lambda a: a[0] in list(stats4.keys()), [(key, value) for key, value in stats3.items()]))
        self.sort_list_of_tuples_by_value(stats3)
        stats3 = dict(stats3[:10])
        stats5 = dict(stats5[:10])

        return stats, vacancies_number, stats2, vacancies_number_by_name, stats3, stats5

    @staticmethod
    def print_statistic(stats1, stats2, stats3, stats4, stats5, stats6):
        """Метод печатает собранную статистику

        Args:
            stats1 dict[int, int]: Динамика уровня зарплат по годам
            stats2 dict[int, int]: Динамика количества вакансий по годам
            stats3 dict[int, int]: Динамика уровня зарплат по годам для выбранной профессии
            stats4 dict[int, int]: Динамика количества вакансий по годам для выбранной профессии
            stats5 dict[str, int]: Уровень зарплат по городам (в порядке убывания)
            stats6 dict[str, int]: Доля вакансий по городам (в порядке убывания)
        """
        print('Динамика уровня зарплат по годам: {0}'.format(stats1))
        print('Динамика количества вакансий по годам: {0}'.format(stats2))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(stats3))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(stats4))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(stats5))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(stats6))


class InputConnect:
    """Класс отвечает за сбор информации из консоли и связывание воедино двух классов: DataSet и Report

    Attributes:
        file_name (str): Название файла или путь до файла + название файла
        vacancy_name (str): Название вакансии для сбора особой статистики
    """
    def __init__(self):
        """Инициализирует класс InputConnect, берёт информацию из консоли, создаёт экземпляр класса,
        собирает статистику, выводит её и позволяет сгенерировать excel-таблицу
        """
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')

        dataset = DataSet(self.file_name, self.vacancy_name)
        stats1, stats2, stats3, stats4, stats5, stats6 = dataset.get_statistic()
        dataset.print_statistic(stats1, stats2, stats3, stats4, stats5, stats6)

        report = Report(self.vacancy_name, stats1, stats2, stats3, stats4, stats5, stats6)
        report.generate_excel()


class Report:
    """Класс отвечает за генерацию excel-файла из собранной статистики

    Attributes:
        wb (Workbook): Экземпляр класса Workbook
        vacancy_name (str): Название вакансии для сбора особой статистики
        stats1: Динамика уровня зарплат по годам
        stats2: Динамика количества вакансий по годам
        stats3: Динамика уровня зарплат по годам для выбранной профессии
        stats4: Динамика количества вакансий по годам для выбранной профессии
        stats5: Уровень зарплат по городам (в порядке убывания)
        stats6: Доля вакансий по городам (в порядке убывания)
    """
    def __init__(self, vacancy_name, stats1, stats2, stats3, stats4, stats5, stats6):
        """Инициализирует класс Report и создаёт экземпляр класс Workbook, отвечающего за создание excel-таблиц

        Args:
            vacancy_name (str): Название вакансии для сбора особой статистики
            stats1: Динамика уровня зарплат по годам
            stats2: Динамика количества вакансий по годам
            stats3: Динамика уровня зарплат по годам для выбранной профессии
            stats4: Динамика количества вакансий по годам для выбранной профессии
            stats5: Уровень зарплат по городам (в порядке убывания)
            stats6: Доля вакансий по городам (в порядке убывания)
        """
        self.wb = Workbook()
        self.vacancy_name = vacancy_name
        self.stats1 = stats1
        self.stats2 = stats2
        self.stats3 = stats3
        self.stats4 = stats4
        self.stats5 = stats5
        self.stats6 = stats6

    def generate_excel(self):
        """ Метод из собранной статистики генерирует файл - report.xlsx,
        где выводится вся собранна статистика в два листа, а также сохраняет файл report.xlsx в директорию,
        откуда запускается этот Python-скрипт
        """
        ws1 = self.wb.active
        ws1.title = 'Статистика по годам'
        ws1.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancy_name, 'Количество вакансий', 'Количество вакансий - ' + self.vacancy_name])
        for year in self.stats1.keys():
            ws1.append([year, self.stats1[year], self.stats3[year], self.stats2[year], self.stats4[year]])

        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancy_name, ' Количество вакансий', ' Количество вакансий - ' + self.vacancy_name]]
        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
            ws1.column_dimensions[get_column_letter(i)].width = column_width + 2

        data = []
        data.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'])
        for (city1, value1), (city2, value2) in zip(self.stats5.items(), self.stats6.items()):
            data.append([city1, value1, '', city2, value2])
        ws2 = self.wb.create_sheet('Статистика по городам')
        for row in data:
            ws2.append(row)

        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                cell = str(cell)
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
            ws2.column_dimensions[get_column_letter(i)].width = column_width + 2

        font_bold = Font(bold=True)
        for col in 'ABCDE':
            ws1[col + '1'].font = font_bold
            ws2[col + '1'].font = font_bold

        for index, _ in enumerate(self.stats5):
            ws2['E' + str(index + 2)].number_format = '0.00%'

        thin = Side(border_style='thin', color='00000000')

        for row in range(len(data)):
            for col in 'ABDE':
                ws2[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

        self.stats1[1] = 1
        for row, _ in enumerate(self.stats1):
            for col in 'ABCDE':
                ws1[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

        self.wb.save('report.xlsx')


if __name__ == '__main__':
    InputConnect()
